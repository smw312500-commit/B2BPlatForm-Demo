# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

def thin():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def create():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '생산등록'

    # 1행: 헤더
    headers = ['라벨코드', '주문량(장)', '납기일']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font      = Font(name='맑은 고딕', size=11, bold=True, color='14532D')
        cell.fill      = PatternFill('solid', fgColor='DCFCE7')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = thin()
    ws.row_dimensions[1].height = 20

    # 2~3행: 샘플
    samples = [
        ['W3MJW01NV', 5000, '2026-06-10'],
        ['W1WTC01BK', 3000, '2026-06-15'],
    ]
    for r, row in enumerate(samples, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font      = Font(name='맑은 고딕', size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = thin()
        ws.row_dimensions[r].height = 18

    # 4~30행: 빈 입력행
    for r in range(4, 31):
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c, value='')
            cell.font      = Font(name='맑은 고딕', size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = thin()
        ws.row_dimensions[r].height = 18

    # 컬럼 너비
    for i, w in enumerate([16, 14, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    out = os.path.join(os.path.dirname(__file__), '생산등록_양식.xlsx')
    wb.save(out)
    print(f'저장: {out}')

if __name__ == '__main__':
    create()
