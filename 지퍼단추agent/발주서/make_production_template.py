import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "생산등록"

# 컬럼: 품목 | 컬러/사이즈 | 주문량(개) | 납기일
cols       = ['품목',       '컬러/사이즈', '주문량(개)', '납기일'   ]
col_widths = [16,            12,            14,           14         ]

for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 22

thin   = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center")

h_fill = PatternFill("solid", fgColor="2D4A6B")
h_font = Font(bold=True, color="FFFFFF", size=10)
for col, name in enumerate(cols, 1):
    c = ws.cell(row=1, column=col, value=name)
    c.fill = h_fill; c.font = h_font; c.alignment = center; c.border = border

# 품목 / 컬러 분리
samples = [
    ("원목단추",      "브라운",  500,  "2026-06-10"),
    ("원목단추",      "블랙",    300,  "2026-06-10"),
    ("플라스틱단추",  "블랙",    2000, "2026-06-12"),
    ("플라스틱단추",  "화이트",  1500, "2026-06-12"),
    ("금속단추",      "실버",    800,  "2026-06-15"),
    ("금속단추",      "블랙",    600,  "2026-06-15"),
    ("지퍼",          "소형",    400,  "2026-06-18"),
    ("지퍼",          "중형",    350,  "2026-06-18"),
    ("지퍼",          "대형",    200,  "2026-06-18"),
]

d_font = Font(size=10)
for row_idx, row_data in enumerate(samples, 2):
    fill = PatternFill("solid", fgColor="EBF4FF" if row_idx % 2 == 0 else "FFFFFF")
    for col_idx, val in enumerate(row_data, 1):
        c = ws.cell(row=row_idx, column=col_idx, value=val)
        c.font = d_font; c.alignment = center; c.border = border; c.fill = fill

# 빈 행 5줄
for row_idx in range(len(samples) + 2, len(samples) + 7):
    fill = PatternFill("solid", fgColor="EBF4FF" if row_idx % 2 == 0 else "FFFFFF")
    for col_idx in range(1, 5):
        c = ws.cell(row=row_idx, column=col_idx)
        c.font = d_font; c.alignment = center; c.border = border; c.fill = fill

out = os.path.join(os.path.dirname(__file__), "생산등록_양식.xlsx")
wb.save(out)
print(f"저장 완료: {out}")
