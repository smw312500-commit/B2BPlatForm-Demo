"""
원자재 발주서 양식 생성 스크립트
실행: python make_order_template.py
결과: 원자재발주_양식.xlsx 생성
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "발주등록"

# ── 헤더 ──
headers = ["품목(원자재명)", "발주처", "발주일", "납기요청일", "수량(kg)", "비고"]
col_widths = [18, 16, 14, 14, 12, 18]

header_fill   = PatternFill("solid", fgColor="4F46E5")
header_font   = Font(bold=True, color="FFFFFF", size=11)
header_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)

for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border    = header_border
    ws.column_dimensions[get_column_letter(col)].width = w

ws.row_dimensions[1].height = 22

# ── 예시 데이터 ──
today = date.today()
sample_data = [
    ["면 원사",      "한국원사공업",   str(today),             str(today + timedelta(days=14)), 500, ""],
    ["폴리에스터 원사", "글로벌섬유",    str(today),             str(today + timedelta(days=19)), 300, "급발주"],
    ["린넨 원사",    "자연섬유사",     str(today + timedelta(1)), str(today + timedelta(days=17)), 200, ""],
    ["울 원사",      "울코리아",       str(today + timedelta(1)), str(today + timedelta(days=20)), 150, ""],
]

sample_fill   = PatternFill("solid", fgColor="EEF2FF")
data_border   = Border(
    left=Side(style='thin', color="D1D5DB"),
    right=Side(style='thin', color="D1D5DB"),
    top=Side(style='thin', color="D1D5DB"),
    bottom=Side(style='thin', color="D1D5DB"),
)

for row_idx, row in enumerate(sample_data, start=2):
    for col_idx, val in enumerate(row, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.fill      = sample_fill
        cell.border    = data_border
        cell.alignment = Alignment(vertical='center')

# ── 빈 입력 행 (10줄) ──
empty_fill = PatternFill("solid", fgColor="FAFAFA")
for row_idx in range(len(sample_data) + 2, len(sample_data) + 12):
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=col_idx, value="")
        cell.border    = data_border
        cell.fill      = empty_fill
        cell.alignment = Alignment(vertical='center')

ws.freeze_panes = "A2"

# ── 코드 안내 시트 ──
note_ws = wb.create_sheet("코드안내")
note_ws.column_dimensions['A'].width = 16
note_ws.column_dimensions['B'].width = 50
notes = [
    ("항목", "설명"),
    ("품목(원자재명)", "면 원사 / 폴리에스터 원사 / 린넨 원사 / 울 원사 / 혼방 원사 등 자유 입력"),
    ("발주처", "공급업체명 자유 입력"),
    ("발주일 / 납기요청일", "YYYY-MM-DD 형식 (예: 2026-06-15)"),
    ("수량(kg)", "숫자만 입력"),
    ("비고", "선택 입력"),
]
for r, (a, b) in enumerate(notes, start=1):
    note_ws.cell(r, 1).value = a
    note_ws.cell(r, 2).value = b
    if r == 1:
        note_ws.cell(r, 1).font = Font(bold=True)
        note_ws.cell(r, 2).font = Font(bold=True)

wb.save("원자재발주_양식.xlsx")
print("✅ 원자재발주_양식.xlsx 생성 완료")
