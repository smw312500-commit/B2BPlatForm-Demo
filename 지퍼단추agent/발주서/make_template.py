import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "원자재발주서"

# 컬럼: 품목 | 발주량 | 발주처 | 발주일 | 납기요청일 | 비고
cols       = ['품목',  '발주량', '발주처', '발주일',     '납기요청일',  '비고']
col_widths = [16,      12,       18,       14,           14,            20   ]

for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 22

thin   = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center")

# 헤더
h_fill = PatternFill("solid", fgColor="2D3748")
h_font = Font(bold=True, color="FFFFFF", size=10)
for col, name in enumerate(cols, 1):
    c = ws.cell(row=1, column=col, value=name)
    c.fill = h_fill; c.font = h_font; c.alignment = center; c.border = border

# 샘플 4종 (날짜는 사용자가 직접 입력)
samples = [
    ("원목",         100, "원목공급사",    "2026-06-01", "2026-06-15", ""),
    ("플라스틱원료", 200, "플라스틱공급사", "2026-06-01", "2026-06-15", ""),
    ("금속원료",     150, "금속공급사",    "2026-06-01", "2026-06-15", ""),
    ("지퍼테이프",   500, "부자재공급사",  "2026-06-01", "2026-06-15", ""),
]

d_font = Font(size=10)
for row_idx, row_data in enumerate(samples, 2):
    fill = PatternFill("solid", fgColor="F7FAFC" if row_idx % 2 == 0 else "FFFFFF")
    for col_idx, val in enumerate(row_data, 1):
        c = ws.cell(row=row_idx, column=col_idx, value=val)
        c.font = d_font; c.alignment = center; c.border = border; c.fill = fill

# 추가 빈 행 10줄
for row_idx in range(6, 16):
    fill = PatternFill("solid", fgColor="F7FAFC" if row_idx % 2 == 0 else "FFFFFF")
    for col_idx in range(1, 7):
        c = ws.cell(row=row_idx, column=col_idx)
        c.font = d_font; c.alignment = center; c.border = border; c.fill = fill

out = os.path.join(os.path.dirname(__file__), "원자재발주서.xlsx")
wb.save(out)
print(f"저장 완료: {out}")
