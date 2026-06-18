"""
생산 등록 양식 생성 스크립트
실행: python make_production_template.py
결과: 생산등록_양식.xlsx 생성
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

FABRIC_NAMES = {
    'C': '면(Cotton)',
    'P': '폴리에스터',
    'L': '린넨(Linen)',
    'W': '울(Wool)',
    'M': '혼방(Mixed)',
}
COLOR_NAMES = {
    'BK': '블랙', 'WH': '화이트', 'NV': '네이비',
    'GY': '그레이', 'BE': '베이지', 'RD': '레드',
}
STAGES = ["원사입고", "정경·제직", "염색", "가공", "검품", "완성"]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "생산등록"

headers    = ["원단코드", "컬러코드", "생산량(야드)", "시작단계", "목표일", "담당자", "비고"]
col_widths = [12, 12, 14, 14, 14, 12, 18]

header_fill = PatternFill("solid", fgColor="059669")
header_font = Font(bold=True, color="FFFFFF", size=11)
thin_border = Border(
    left=Side(style='thin', color="D1D5DB"),
    right=Side(style='thin', color="D1D5DB"),
    top=Side(style='thin', color="D1D5DB"),
    bottom=Side(style='thin', color="D1D5DB"),
)

for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border    = thin_border
    ws.column_dimensions[get_column_letter(col)].width = w

ws.row_dimensions[1].height = 22

today = date.today()
sample_data = [
    ["C", "NV", 500, "원사입고",  str(today + timedelta(days=14)), "김생산", ""],
    ["P", "BK", 300, "정경·제직", str(today + timedelta(days=10)), "이공정", "긴급"],
    ["L", "WH", 200, "원사입고",  str(today + timedelta(days=18)), "박직조", ""],
    ["W", "GY", 150, "원사입고",  str(today + timedelta(days=20)), "최염색", ""],
    ["M", "BE", 250, "원사입고",  str(today + timedelta(days=22)), "",       ""],
]

sample_fill = PatternFill("solid", fgColor="ECFDF5")
for row_idx, row in enumerate(sample_data, start=2):
    for col_idx, val in enumerate(row, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.fill      = sample_fill
        cell.border    = thin_border
        cell.alignment = Alignment(vertical='center')

empty_fill = PatternFill("solid", fgColor="FAFAFA")
for row_idx in range(len(sample_data) + 2, len(sample_data) + 15):
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=col_idx, value="")
        cell.border    = thin_border
        cell.fill      = empty_fill
        cell.alignment = Alignment(vertical='center')

ws.freeze_panes = "A2"

# ── 코드 안내 시트 ──
note_ws = wb.create_sheet("코드안내")
note_ws.column_dimensions['A'].width = 16
note_ws.column_dimensions['B'].width = 55

notes = [
    ("항목",        "설명 / 유효값"),
    ("원단코드",    "C=면(Cotton) / P=폴리에스터 / L=린넨(Linen) / W=울(Wool) / M=혼방(Mixed)"),
    ("컬러코드",    "BK=블랙 / WH=화이트 / NV=네이비 / GY=그레이 / BE=베이지 / RD=레드"),
    ("생산량(야드)", "숫자만 입력 (소수점 가능)"),
    ("시작단계",    " / ".join(STAGES)),
    ("목표일",      "YYYY-MM-DD 형식 (예: 2026-07-15)"),
    ("담당자",      "선택 입력"),
    ("비고",        "선택 입력"),
]

for r, (a, b) in enumerate(notes, start=1):
    note_ws.cell(r, 1).value = a
    note_ws.cell(r, 2).value = b
    if r == 1:
        note_ws.cell(r, 1).font = Font(bold=True)
        note_ws.cell(r, 2).font = Font(bold=True)

wb.save("생산등록_양식.xlsx")
print("✅ 생산등록_양식.xlsx 생성 완료")
