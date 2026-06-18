"""
발주서 엑셀 양식 생성 - 스크린샷 기준 심플 구조
행1: 헤더 / 행2~3: 샘플 / 행4~: 빈칸
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

def thin():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def create_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "발주등록"

    # ── 1행: 헤더 ─────────────────────────────
    headers = ["품목", "발주량", "발주처", "발주일", "납기요청일", "비고"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font      = Font(name="맑은 고딕", size=11, bold=True, color="1E3A8A")
        cell.fill      = PatternFill("solid", fgColor="DBEAFE")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin()
    ws.row_dimensions[1].height = 20

    # ── 2~3행: 샘플 데이터 ───────────────────
    samples = [
        ["라벨원단", 100, "원단공급사A", "2026-05-22", "2026-06-10", ""],
        ["잉크",       5, "잉크공급사B", "2026-05-22", "2026-06-10", "샘플 발주"],
    ]
    for r, row in enumerate(samples, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font      = Font(name="맑은 고딕", size=11)
            cell.alignment = Alignment(horizontal="center" if c in (1,2,4,5) else "left", vertical="center")
            cell.border    = thin()
        ws.row_dimensions[r].height = 18

    # ── 4~30행: 빈 입력행 ────────────────────
    for r in range(4, 31):
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c, value="")
            cell.font      = Font(name="맑은 고딕", size=11)
            cell.alignment = Alignment(horizontal="center" if c in (1,2,4,5) else "left", vertical="center")
            cell.border    = thin()
        ws.row_dimensions[r].height = 18

    # ── 품목 드롭다운 (A2:A30) ───────────────
    dv = DataValidation(type="list", formula1='"라벨원단,잉크"', allow_blank=True)
    dv.sqref = "A2:A30"
    ws.add_data_validation(dv)

    # ── 컬럼 너비 ─────────────────────────────
    for i, w in enumerate([14, 10, 18, 14, 14, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    out = os.path.join(os.path.dirname(__file__), "발주서_양식.xlsx")
    wb.save(out)
    print(f"저장: {out}")

if __name__ == "__main__":
    create_template()
